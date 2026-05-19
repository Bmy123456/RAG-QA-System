import openpyxl


def parse_xlsx(file_path: str) -> str:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            row_str = " | ".join(str(cell) if cell is not None else "" for cell in row)
            if row_str.strip():
                rows.append(row_str)
        parts.append(f"[Sheet: {sheet_name}]\n" + "\n".join(rows))
    return "\n\n".join(parts)
